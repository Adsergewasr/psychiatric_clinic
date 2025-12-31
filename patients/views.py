from django.db import models
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.http import require_POST
import json
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Patient, Diagnosis
from .forms import PatientForm, PatientSearchForm, PatientExportForm


@login_required
def patient_list(request):
    """Список пациентов с поиском и фильтрацией"""
    form = PatientSearchForm(request.GET or None)
    patients = Patient.objects.all().select_related('attending_physician')
    
    # Применяем фильтры
    if form.is_valid():
        query = form.cleaned_data.get('query')
        status = form.cleaned_data.get('status')
        gender = form.cleaned_data.get('gender')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        
        if query:
            patients = patients.filter(
                Q(last_name__icontains=query) |
                Q(first_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(case_number__icontains=query) |
                Q(passport_series__icontains=query) |
                Q(passport_number__icontains=query) |
                Q(snils__icontains=query) |
                Q(phone__icontains=query) |
                Q(address__icontains=query)
            )
        
        if status:
            patients = patients.filter(status=status)
        
        if gender:
            patients = patients.filter(gender=gender)
        
        if date_from:
            patients = patients.filter(admission_date__date__gte=date_from)
        
        if date_to:
            patients = patients.filter(admission_date__date__lte=date_to)
    
    # Сортировка
    sort_by = request.GET.get('sort', '-admission_date')
    if sort_by.lstrip('-') in ['last_name', 'first_name', 'admission_date', 'birth_date']:
        patients = patients.order_by(sort_by)
    
    # Пагинация
    paginator = Paginator(patients, 25)  # 25 пациентов на странице
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_patients': patients.count(),
        'sort_by': sort_by,
        'status_counts': {
            'HOSPITALIZED': Patient.objects.filter(status='HOSPITALIZED').count(),
            'DISCHARGED': Patient.objects.filter(status='DISCHARGED').count(),
            'TRANSFERRED': Patient.objects.filter(status='TRANSFERRED').count(),
            'DIED': Patient.objects.filter(status='DIED').count(),
        }
    }
    
    return render(request, 'patients/patient_list.html', context)


@login_required
@permission_required('patients.add_patient', raise_exception=True)
def patient_create(request):
    """Создание нового пациента"""
    if request.method == 'POST':
        form = PatientForm(request.POST, user=request.user)
        if form.is_valid():
            patient = form.save()
            messages.success(
                request,
                f'Пациент {patient.full_name} успешно создан. '
                f'Номер истории болезни: {patient.case_number}'
            )
            return redirect('patients:patient_detail', pk=patient.pk)
    else:
        form = PatientForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Создание новой карты пациента',
        'submit_text': 'Создать карту',
    }
    
    return render(request, 'patients/patient_form.html', context)


@login_required
def patient_detail(request, pk):
    """Просмотр карты пациента"""
    patient = get_object_or_404(
        Patient.objects.select_related('attending_physician', 'created_by'),
        pk=pk
    )
    
    # Проверка прав доступа
    if not request.user.has_perm('patients.view_patient'):
        if patient.attending_physician != request.user:
            messages.error(request, 'У вас нет прав для просмотра этой карты')
            return redirect('patients:patient_list')
    
    # Получаем историю госпитализаций
    hospitalizations = patient.hospitalizations.all()
    
    # Подготовим данные для шаблона
    context = {
        'patient': patient,
        'hospitalizations': hospitalizations,
        'can_edit': request.user.has_perm('patients.change_patient') or 
                    (patient.attending_physician and patient.attending_physician == request.user),
        'can_delete': request.user.has_perm('patients.delete_patient'),
    }
    
    # Добавим безопасные атрибуты
    if patient.attending_physician:
        context['attending_physician_name'] = patient.attending_physician.get_full_name() or patient.attending_physician.username
    else:
        context['attending_physician_name'] = "Не назначен"
    
    if patient.created_by:
        context['created_by_name'] = patient.created_by.get_full_name() or patient.created_by.username
    else:
        context['created_by_name'] = "Система"
    
    return render(request, 'patients/patient_detail.html', context)

@login_required
def patient_update(request, pk):
    """Редактирование карты пациента"""
    patient = get_object_or_404(Patient, pk=pk)
    
    # Проверка прав доступа
    if not request.user.has_perm('patients.change_patient'):
        if patient.attending_physician != request.user:
            messages.error(request, 'У вас нет прав для редактирования этой карты')
            return redirect('patients:patient_detail', pk=pk)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient, user=request.user)
        if form.is_valid():
            patient = form.save()
            messages.success(request, f'Карта пациента {patient.full_name} успешно обновлена')
            return redirect('patients:patient_detail', pk=patient.pk)
    else:
        form = PatientForm(instance=patient, user=request.user)
    
    context = {
        'form': form,
        'patient': patient,
        'title': f'Редактирование карты пациента: {patient.full_name}',
        'submit_text': 'Сохранить изменения',
    }
    
    return render(request, 'patients/patient_form.html', context)


@login_required
@permission_required('patients.delete_patient', raise_exception=True)
def patient_delete(request, pk):
    """Удаление карты пациента"""
    patient = get_object_or_404(Patient, pk=pk)
    
    if request.method == 'POST':
        patient_name = patient.full_name
        case_number = patient.case_number
        patient.delete()
        messages.success(
            request,
            f'Карта пациента {patient_name} (ИБ: {case_number}) удалена'
        )
        return redirect('patients:patient_list')
    
    context = {
        'patient': patient,
        'title': 'Подтверждение удаления',
    }
    
    return render(request, 'patients/patient_confirm_delete.html', context)


@login_required
@require_POST
def patient_discharge(request, pk):
    """Быстрая выписка пациента"""
    patient = get_object_or_404(Patient, pk=pk)
    
    # Проверка прав доступа
    if not request.user.has_perm('patients.change_patient'):
        if patient.attending_physician != request.user:
            messages.error(request, 'У вас нет прав для выписки этого пациента')
            return redirect('patients:patient_detail', pk=pk)
    
    if patient.status == 'HOSPITALIZED':
        patient.status = 'DISCHARGED'
        patient.discharge_date = timezone.now()
        patient.save()
        messages.success(request, f'Пациент {patient.full_name} выписан')
    else:
        messages.warning(request, f'Пациент уже имеет статус: {patient.get_status_display()}')
    
    return redirect('patients:patient_detail', pk=pk)


@login_required
def patient_export(request):
    """Экспорт пациентов"""
    if request.method == 'POST':
        form = PatientExportForm(request.POST)
        if form.is_valid():
            patients = form.cleaned_data['patients']
            export_format = form.cleaned_data['export_format']
            include_fields = form.cleaned_data.get('include_fields', [])
            
            # Подготовка данных для экспорта
            data = prepare_export_data(patients, include_fields)
            
            if export_format == 'csv':
                return export_csv(data, patients)
            elif export_format == 'xlsx':
                return export_excel(data, patients)
            elif export_format == 'json':
                return export_json(data, patients)
    else:
        # По умолчанию выбираем всех пациентов
        patients = Patient.objects.all()
        initial = {
            'patients': patients,
            'include_fields': ['basic', 'documents', 'hospitalization'],
        }
        form = PatientExportForm(initial=initial)
    
    context = {
        'form': form,
        'total_patients': Patient.objects.count(),
    }
    
    return render(request, 'patients/patient_export.html', context)


def prepare_export_data(patients, include_fields):
    """Подготовка данных для экспорта"""
    data = []
    
    for patient in patients:
        row = {}
        
        if 'basic' in include_fields:
            row.update({
                'Номер истории болезни': patient.case_number,
                'Фамилия': patient.last_name,
                'Имя': patient.first_name,
                'Отчество': patient.middle_name,
                'Пол': patient.get_gender_display(),
                'Дата рождения': patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '',
                'Возраст': patient.age,
                'Место рождения': patient.birth_place,
                'Гражданство': patient.citizenship,
                'Адрес': patient.address,
                'Телефон': patient.phone,
            })
        
        if 'documents' in include_fields:
            row.update({
                'Серия паспорта': patient.passport_series,
                'Номер паспорта': patient.passport_number,
                'Кем выдан': patient.passport_issued_by,
                'Дата выдачи': patient.passport_issue_date.strftime('%d.%m.%Y') if patient.passport_issue_date else '',
                'СНИЛС': patient.snils,
                'Страховой полис': patient.insurance_policy,
            })
        
        if 'hospitalization' in include_fields:
            row.update({
                'Дата поступления': patient.admission_date.strftime('%d.%m.%Y %H:%M') if patient.admission_date else '',
                'Откуда поступил': patient.admission_from,
                'Кем доставлен': patient.delivered_by,
                'Диагноз направившего учреждения': patient.referral_diagnosis,
                'Диагноз при поступлении': patient.admission_diagnosis,
                'Код МКБ при поступлении': patient.admission_mkb_code,
                'Лечащий врач': patient.attending_physician.get_full_name() if patient.attending_physician else '',
            })
        
        if 'discharge' in include_fields:
            row.update({
                'Дата выписки': patient.discharge_date.strftime('%d.%m.%Y %H:%M') if patient.discharge_date else '',
                'Диагноз при выписке': patient.discharge_diagnosis,
                'Код МКБ при выписке': patient.discharge_mkb_code,
                'Исход заболевания': patient.get_outcome_display() if patient.outcome else '',
                'Трудоспособность': patient.get_work_capacity_display() if patient.work_capacity else '',
                'Статус': patient.get_status_display(),
            })
        
        if 'notes' in include_fields:
            row.update({
                'Примечания': patient.notes,
                'Дата создания': patient.created_at.strftime('%d.%m.%Y %H:%M') if patient.created_at else '',
                'Создал': patient.created_by.get_full_name() if patient.created_by else '',
            })
        
        data.append(row)
    
    return data


def export_csv(data, patients):
    """Экспорт в CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    if data:
        writer = csv.DictWriter(response, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    return response


def export_excel(data, patients):
    """Экспорт в Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пациенты"
    
    if data:
        # Заголовки
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws.column_dimensions[col_letter].width = 20
        
        # Данные
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = row_data.get(header, '')
    
    wb.save(response)
    return response


def export_json(data, patients):
    """Экспорт в JSON"""
    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json"'
    return response


@login_required
def api_diagnoses(request):
    """API для автодополнения диагнозов"""
    query = request.GET.get('q', '')
    
    if query:
        diagnoses = Diagnosis.objects.filter(
            Q(code__icontains=query) | Q(name__icontains=query)
        ).order_by('code')[:10]
    else:
        diagnoses = Diagnosis.objects.none()
    
    results = [
        {
            'id': d.code,
            'text': f'{d.code} - {d.name}',
            'code': d.code,
            'name': d.name,
            'description': d.description[:100] if d.description else '',
        }
        for d in diagnoses
    ]
    
    return JsonResponse({'results': results})


@login_required
def dashboard(request):
    """Дашборд с общей статистикой"""
    # Основная статистика
    total_patients = Patient.objects.count()
    hospitalized = Patient.objects.filter(status='HOSPITALIZED').count()
    discharged = Patient.objects.filter(status='DISCHARGED').count()
    
    # Статистика за последние 30 дней
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
    recent_admissions = Patient.objects.filter(
        admission_date__gte=thirty_days_ago
    ).count()
    
    recent_discharges = Patient.objects.filter(
        discharge_date__gte=thirty_days_ago
    ).count()
    
    # Распределение по полу
    gender_stats = Patient.objects.values('gender').annotate(
        count=models.Count('id')
    )
    
    # Распределение по возрасту
    age_groups = {
        'До 18 лет': Patient.objects.filter(birth_date__gte=timezone.now() - timezone.timedelta(days=18*365)).count(),
        '18-30 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=18*365),
            birth_date__gte=timezone.now() - timezone.timedelta(days=30*365)
        ).count(),
        '31-50 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=30*365),
            birth_date__gte=timezone.now() - timezone.timedelta(days=50*365)
        ).count(),
        '51-70 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=50*365),
            birth_date__gte=timezone.now() - timezone.timedelta(days=70*365)
        ).count(),
        'Старше 70 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=70*365)
        ).count(),
    }
    
    # Последние поступления
    recent_patients = Patient.objects.select_related('attending_physician').order_by('-admission_date')[:10]
    
    context = {
        'total_patients': total_patients,
        'hospitalized': hospitalized,
        'discharged': discharged,
        'recent_admissions': recent_admissions,
        'recent_discharges': recent_discharges,
        'gender_stats': gender_stats,
        'age_groups': age_groups,
        'recent_patients': recent_patients,
    }
    
    return render(request, 'patients/dashboard.html', context)