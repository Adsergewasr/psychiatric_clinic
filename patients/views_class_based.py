from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView, TemplateView, View
from django.views.generic.edit import FormMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.db.models import Q, Count
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import models
import json
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from users.mixins import RoleRequiredMixin, ObjectPermissionMixin
from .models import Patient, Diagnosis, Hospitalization
from .forms import PatientForm, PatientSearchForm, PatientExportForm


class DashboardView(RoleRequiredMixin, TemplateView):
    """Дашборд с общей статистикой (классовое представление)"""
    template_name = 'patients/dashboard.html'
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR', 'ANALYST']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Основная статистика
        context['total_patients'] = Patient.objects.count()
        context['hospitalized'] = Patient.objects.filter(status='HOSPITALIZED').count()
        context['discharged'] = Patient.objects.filter(status='DISCHARGED').count()
        
        # Статистика за последние 30 дней
        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
        context['recent_admissions'] = Patient.objects.filter(
            admission_date__gte=thirty_days_ago
        ).count()
        
        context['recent_discharges'] = Patient.objects.filter(
            discharge_date__gte=thirty_days_ago
        ).count()
        
        # Распределение по полу
        context['gender_stats'] = Patient.objects.values('gender').annotate(
            count=Count('id')
        )
        
        # Распределение по возрасту
        today = timezone.now().date()
        context['age_groups'] = {
            'До 18 лет': Patient.objects.filter(
                birth_date__gte=today.replace(year=today.year - 18)
            ).count(),
            '18-30 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 18),
                birth_date__gte=today.replace(year=today.year - 30)
            ).count(),
            '31-50 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 30),
                birth_date__gte=today.replace(year=today.year - 50)
            ).count(),
            '51-70 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 50),
                birth_date__gte=today.replace(year=today.year - 70)
            ).count(),
            'Старше 70 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 70)
            ).count(),
        }
        
        # Последние поступления
        context['recent_patients'] = Patient.objects.select_related(
            'attending_physician'
        ).order_by('-admission_date')[:10]
        
        # Статистика по врачам (только для администраторов и аналитиков)
        if self.request.user.is_administrator or self.request.user.is_analyst:
            context['doctor_stats'] = Patient.objects.filter(
                attending_physician__isnull=False
            ).values(
                'attending_physician__last_name',
                'attending_physician__first_name'
            ).annotate(
                patient_count=Count('id'),
                hospitalized_count=Count('id', filter=Q(status='HOSPITALIZED'))
            ).order_by('-patient_count')[:5]
        
        # Права доступа
        context['can_export'] = (
            self.request.user.is_administrator or
            self.request.user.is_analyst or
            self.request.user.has_perm('patients.export_patients')
        )
        context['can_create'] = (
            self.request.user.is_administrator or
            self.request.user.is_doctor or
            self.request.user.is_nurse or
            self.request.user.is_registrar
        )
        
        return context


class PatientListView(RoleRequiredMixin, ListView):
    """Список пациентов с поиском и фильтрацией (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_list.html'
    context_object_name = 'patients_with_perms'
    paginate_by = 25
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR', 'ANALYST']
    
    def get_queryset(self):
        # Базовый queryset с учетом прав доступа
        if self.request.user.is_administrator or self.request.user.has_perm('patients.view_all_patients'):
            queryset = Patient.objects.all()
        elif self.request.user.is_doctor:
            queryset = Patient.objects.filter(attending_physician=self.request.user)
        else:
            queryset = Patient.objects.all()
        
        queryset = queryset.select_related('attending_physician')
        
        # Применяем фильтры из формы поиска
        self.search_form = PatientSearchForm(self.request.GET or None)
        if self.search_form.is_valid():
            query = self.search_form.cleaned_data.get('query')
            status = self.search_form.cleaned_data.get('status')
            gender = self.search_form.cleaned_data.get('gender')
            date_from = self.search_form.cleaned_data.get('date_from')
            date_to = self.search_form.cleaned_data.get('date_to')
            
            if query:
                queryset = queryset.filter(
                    Q(last_name__icontains=query) |
                    Q(first_name__icontains=query) |
                    Q(middle_name__icontains=query) |
                    Q(case_number__icontains=query) |
                    Q(passport_series__icontains=query) |
                    Q(passport_number__icontains=query) |
                    Q(inn__icontains=query) |
                    Q(phone__icontains=query) |
                    Q(address__icontains=query)
                )
            if status:
                queryset = queryset.filter(status=status)
            if gender:
                queryset = queryset.filter(gender=gender)
            if date_from:
                queryset = queryset.filter(admission_date__date__gte=date_from)
            if date_to:
                queryset = queryset.filter(admission_date__date__lte=date_to)
        
        # Сортировка
        sort_by = self.request.GET.get('sort', '-admission_date')
        if sort_by in ['last_name', 'admission_date', 'birth_date', 'case_number']:
            queryset = queryset.order_by(sort_by)
        elif sort_by == '-last_name':
            queryset = queryset.order_by('-last_name')
        else:
            queryset = queryset.order_by('-admission_date')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика по статусам с учетом прав доступа
        status_counts = {}
        for status_code, status_name in Patient.STATUS_CHOICES:
            if self.request.user.is_administrator or self.request.user.has_perm('patients.view_all_patients'):
                status_counts[status_code] = Patient.objects.filter(status=status_code).count()
            elif self.request.user.is_doctor:
                status_counts[status_code] = Patient.objects.filter(
                    status=status_code,
                    attending_physician=self.request.user
                ).count()
            else:
                status_counts[status_code] = Patient.objects.filter(status=status_code).count()
        
        # Для каждого пациента вычисляем права
        patient_permissions = []
        for patient in context['object_list']:
            can_edit = patient.user_can_edit(self.request.user)
            can_delete = patient.user_can_delete(self.request.user)
            can_discharge = (patient.status == 'HOSPITALIZED' and can_edit)
            patient_permissions.append({
                'patient': patient,
                'can_edit': can_edit,
                'can_delete': can_delete,
                'can_discharge': can_discharge,
            })
        
        context.update({
            'patients_with_perms': patient_permissions,
            'form': self.search_form if hasattr(self, 'search_form') else PatientSearchForm(),
            'total_patients': self.get_queryset().count(),
            'status_counts': status_counts,
            'sort_by': self.request.GET.get('sort', '-admission_date'),
            'can_create_patient': (
                self.request.user.is_administrator or 
                self.request.user.is_doctor or 
                self.request.user.is_nurse or 
                self.request.user.is_registrar
            ),
        })
        return context


class PatientDetailView(ObjectPermissionMixin, DetailView):
    """Просмотр карты пациента (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_detail.html'
    context_object_name = 'patient'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('attending_physician', 'created_by')
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient = self.get_object()
        
        # Получаем историю госпитализаций
        hospitalizations = patient.hospitalizations.all()
        
        # Определяем доступные действия
        can_edit = patient.user_can_edit(self.request.user)
        can_delete = patient.user_can_delete(self.request.user)
        can_discharge = (patient.status == 'HOSPITALIZED' and can_edit)
        
        context.update({
            'hospitalizations': hospitalizations,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'can_discharge': can_discharge,
        })
        return context
    
    def dispatch(self, request, *args, **kwargs):
        # Проверка прав на просмотр
        patient = self.get_object()
        if not patient.user_can_view(request.user):
            messages.error(request, 'У вас нет прав для просмотра этой карты')
            return redirect('patients:patient_list')
        return super().dispatch(request, *args, **kwargs)


class PatientCreateView(RoleRequiredMixin, CreateView):
    """Создание нового пациента (классовое представление)"""
    model = Patient
    form_class = PatientForm
    template_name = 'patients/patient_form.html'
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR']
    success_url = reverse_lazy('patients:patient_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Пациент успешно добавлен')
        return response
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.pk})


class PatientUpdateView(ObjectPermissionMixin, UpdateView):
    """Редактирование пациента (классовое представление)"""
    model = Patient
    form_class = PatientForm
    template_name = 'patients/patient_form.html'
    context_object_name = 'patient'
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.pk})
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Данные пациента обновлены')
        return response


class PatientDeleteView(ObjectPermissionMixin, DeleteView):
    """Удаление пациента (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_confirm_delete.html'
    success_url = reverse_lazy('patients:patient_list')
    
    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(self.request, 'Пациент удалён')
        return response


class PatientDischargeView(ObjectPermissionMixin, UpdateView):
    """Быстрая выписка пациента (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_discharge.html'
    fields = ['discharge_date', 'discharge_diagnosis', 'discharge_mkb_code', 'outcome', 'work_capacity']
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.pk})
    
    def dispatch(self, request, *args, **kwargs):
        patient = self.get_object()
        # Проверка прав на выписку через user_can_edit
        if not (patient.status == 'HOSPITALIZED' and patient.user_can_edit(request.user)):
            messages.error(request, 'У вас нет прав для выписки этого пациента')
            return redirect('patients:patient_detail', pk=patient.pk)
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        patient = form.save(commit=False)
        patient.status = 'DISCHARGED'
        if not patient.discharge_date:
            patient.discharge_date = timezone.now()
        patient.save()
        messages.success(self.request, f'Пациент {patient.full_name} выписан')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.get_object()
        return context


class PatientQuickDischargeView(ObjectPermissionMixin, View):
    """Быстрая выписка пациента без формы (классовое представление)"""
    
    def post(self, request, pk):
        patient = get_object_or_404(Patient, pk=pk)
        
        # Проверка прав на выписку
        if not (patient.status == 'HOSPITALIZED' and patient.user_can_edit(request.user)):
            messages.error(request, 'У вас нет прав для выписки этого пациента')
            return redirect('patients:patient_detail', pk=pk)
        
        patient.status = 'DISCHARGED'
        patient.discharge_date = timezone.now()
        patient.save()
        
        messages.success(request, f'Пациент {patient.full_name} выписан')
        return redirect('patients:patient_detail', pk=pk)


class PatientExportView(RoleRequiredMixin, FormView):
    """Экспорт пациентов (классовое представление)"""
    template_name = 'patients/patient_export.html'
    form_class = PatientExportForm
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR', 'ANALYST']
    
    def get_initial(self):
        initial = super().get_initial()
        
        # Определяем доступный queryset пациентов
        patients_queryset = self._get_patients_queryset()
        
        initial.update({
            'patients': patients_queryset,
            'include_fields': ['basic', 'documents', 'hospitalization'],
            'export_format': 'xlsx',
        })
        return initial
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        
        # Передаем queryset для поля patients
        if 'data' not in kwargs:
            kwargs['initial'] = self.get_initial()
        
        # Обновляем queryset для поля patients
        form = PatientExportForm(**kwargs)
        form.fields['patients'].queryset = self._get_patients_queryset()
        
        return kwargs
    
    def _get_patients_queryset(self):
        """Получаем queryset пациентов с учетом прав доступа"""
        if self.request.user.is_administrator or self.request.user.has_perm('patients.view_all_patients'):
            return Patient.objects.all()
        elif self.request.user.is_doctor:
            return Patient.objects.filter(attending_physician=self.request.user)
        else:
            return Patient.objects.all()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patients_queryset = self._get_patients_queryset()
        context['total_patients'] = patients_queryset.count()
        return context
    
    def form_valid(self, form):
        patients = form.cleaned_data['patients']
        export_format = form.cleaned_data['export_format']
        include_fields = form.cleaned_data.get('include_fields', [])
        
        # Проверка: пользователь может экспортировать только тех пациентов, которых может просматривать
        for patient in patients:
            if not patient.user_can_view(self.request.user):
                messages.error(self.request, 'Вы можете экспортировать только доступных вам пациентов')
                return self.form_invalid(form)
        
        # Подготовка данных для экспорта
        data = self._prepare_export_data(patients, include_fields)
        
        if export_format == 'csv':
            return self._export_csv(data)
        elif export_format == 'xlsx':
            return self._export_excel(data)
        elif export_format == 'json':
            return self._export_json(data)
        
        return super().form_valid(form)
    
    def _prepare_export_data(self, patients, include_fields):
        """Подготовка данных для экспорта"""
        data = []
        
        for patient in patients:
            row = {}
            
            if 'basic' in include_fields:
                row.update({
                    'Номер истории болезни': patient.case_number,
                    'Фамилия': patient.last_name,
                    'Имя': patient.first_name,
                    'Отчество': patient.middle_name or '',
                    'Пол': patient.get_gender_display(),
                    'Дата рождения': patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '',
                    'Возраст': patient.age,
                    'Место рождения': patient.birth_place or '',
                    'Гражданство': patient.citizenship or 'РФ',
                    'Адрес': patient.address,
                    'Телефон': patient.phone or '',
                })
            
            if 'documents' in include_fields:
                row.update({
                    'Серия паспорта': patient.passport_series or '',
                    'Номер паспорта': patient.passport_number or '',
                    'Кем выдан': patient.passport_issued_by or '',
                    'Дата выдачи': patient.passport_issue_date.strftime('%d.%m.%Y') if patient.passport_issue_date else '',
                    'ИНН': patient.inn or '',
                    'Страховой полис': patient.insurance_policy or '',
                })
            
            if 'hospitalization' in include_fields:
                row.update({
                    'Дата поступления': patient.admission_date.strftime('%d.%m.%Y %H:%M') if patient.admission_date else '',
                    'Откуда поступил': patient.admission_from or '',
                    'Кем доставлен': patient.delivered_by or '',
                    'Диагноз направившего учреждения': patient.referral_diagnosis or '',
                    'Диагноз при поступлении': patient.admission_diagnosis,
                    'Код МКБ при поступлении': patient.admission_mkb_code or '',
                    'Лечащий врач': patient.attending_physician.get_full_name() if patient.attending_physician else '',
                })
            
            if 'discharge' in include_fields:
                row.update({
                    'Дата выписки': patient.discharge_date.strftime('%d.%m.%Y %H:%M') if patient.discharge_date else '',
                    'Диагноз при выписке': patient.discharge_diagnosis or '',
                    'Код МКБ при выписке': patient.discharge_mkb_code or '',
                    'Исход заболевания': patient.get_outcome_display() if patient.outcome else '',
                    'Трудоспособность': patient.get_work_capacity_display() if patient.work_capacity else '',
                    'Статус': patient.get_status_display(),
                })
            
            if 'notes' in include_fields:
                row.update({
                    'Примечания': patient.notes or '',
                    'Дата создания': patient.created_at.strftime('%d.%m.%Y %H:%M') if patient.created_at else '',
                    'Создал': patient.created_by.get_full_name() if patient.created_by else '',
                })
            
            data.append(row)
        
        return data
    
    def _export_csv(self, data):
        """Экспорт в CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        if data:
            writer = csv.DictWriter(response, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return response
    
    def _export_excel(self, data):
        """Экспорт в Excel"""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Пациенты"
        
        if data:
            # Заголовки
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = header
                ws.column_dimensions[col_letter].width = 20
            
            # Данные
            for row_num, row_data in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_num)
                    ws[f'{col_letter}{row_num}'] = row_data.get(header, '')
        
        wb.save(response)
        return response
    
    def _export_json(self, data):
        """Экспорт в JSON"""
        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json"'
        return response


class ApiDiagnosesView(LoginRequiredMixin, View):
    """API для автодополнения диагнозов (классовое представление)"""
    
    def get(self, request):
        query = request.GET.get('q', '')
        
        if query:
            diagnoses = Diagnosis.objects.filter(
                Q(code__icontains=query) | Q(name__icontains=query)
            ).order_by('code')[:10]
        else:
            diagnoses = Diagnosis.objects.none()
        
        results = [
            {
                'id': d.code,
                'text': f'{d.code} - {d.name}',
                'code': d.code,
                'name': d.name,
                'description': d.description[:100] if d.description else '',
            }
            for d in diagnoses
        ]
        
        return JsonResponse({'results': results})


class PatientStatisticsView(RoleRequiredMixin, TemplateView):
    """Расширенная статистика пациентов (классовое представление)"""
    template_name = 'patients/statistics.html'
    allowed_roles = ['ADMIN', 'DOCTOR', 'ANALYST']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Фильтры
        year = self.request.GET.get('year', timezone.now().year)
        month = self.request.GET.get('month')
        
        # Базовый queryset
        queryset = Patient.objects.filter(
            admission_date__year=year
        )
        
        if month:
            queryset = queryset.filter(admission_date__month=month)
        
        # Общая статистика
        context['total_patients'] = queryset.count()
        context['hospitalized'] = queryset.filter(status='HOSPITALIZED').count()
        context['discharged'] = queryset.filter(status='DISCHARGED').count()
        context['transferred'] = queryset.filter(status='TRANSFERRED').count()
        context['died'] = queryset.filter(status='DIED').count()
        
        # Статистика по месяцам
        monthly_stats = []
        for month_num in range(1, 13):
            month_data = queryset.filter(admission_date__month=month_num)
            monthly_stats.append({
                'month': month_num,
                'admissions': month_data.count(),
                'discharges': month_data.filter(discharge_date__month=month_num).count(),
            })
        context['monthly_stats'] = monthly_stats
        
        # Статистика по диагнозам (топ-10)
        context['top_diagnoses'] = Diagnosis.objects.annotate(
            patient_count=Count('patient')
        ).order_by('-patient_count')[:10]
        
        # Статистика по возрасту
        age_stats = {
            'До 18 лет': queryset.filter(age__lt=18).count(),
            '18-30 лет': queryset.filter(age__range=(18, 30)).count(),
            '31-50 лет': queryset.filter(age__range=(31, 50)).count(),
            '51-70 лет': queryset.filter(age__range=(51, 70)).count(),
            'Старше 70 лет': queryset.filter(age__gt=70).count(),
        }
        context['age_stats'] = age_stats
        
        # Годы для фильтра
        context['years'] = Patient.objects.dates('admission_date', 'year')
        
        return context


class HospitalizationCreateView(RoleRequiredMixin, CreateView):
    """Добавление госпитализации (классовое представление)"""
    model = Hospitalization
    template_name = 'patients/hospitalization_form.html'
    fields = ['admission_date', 'diagnosis', 'mkb_code', 'department', 'attending_physician', 'notes']
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE']
    
    def dispatch(self, request, *args, **kwargs):
        self.patient = get_object_or_404(Patient, pk=kwargs['patient_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.patient
        return context
    
    def form_valid(self, form):
        hospitalization = form.save(commit=False)
        hospitalization.patient = self.patient
        hospitalization.save()
        messages.success(self.request, 'Госпитализация добавлена')
        return redirect('patients:patient_detail', pk=self.patient.pk)


class HospitalizationUpdateView(ObjectPermissionMixin, UpdateView):
    """Редактирование госпитализации (классовое представление)"""
    model = Hospitalization
    template_name = 'patients/hospitalization_form.html'
    fields = ['admission_date', 'discharge_date', 'diagnosis', 'mkb_code', 'department', 
              'attending_physician', 'outcome', 'notes']
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.patient.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Госпитализация обновлена')
        return response


class HospitalizationDeleteView(ObjectPermissionMixin, DeleteView):
    """Удаление госпитализации (классовое представление)"""
    model = Hospitalization
    template_name = 'patients/hospitalization_confirm_delete.html'
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.patient.pk})
    
    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(self.request, 'Госпитализация удалена')
        return response